"""Script para exportar el formato excel de solicitudes de ingreso."""
from copy import copy
import os
import shutil
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from sqlalchemy.orm import Session, selectinload

from app.models.branches import Branch, BranchTypes
from app.models.users import Guest, User
from app.models.entrances import EntranceRequest, EntranceRequestGuest


def copy_row(ws, source_row, target_row):
    """Copia una fila de un worksheet a otra fila, incluyendo estilos y comentarios."""
    for col in range(1, ws.max_column + 1):
        source_cell = ws.cell(row=source_row, column=col)
        target_cell = ws.cell(row=target_row, column=col)
        # Copiar valor
        target_cell.value = source_cell.value
        # Copiar estilo
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)
    # Copiar altura de fila
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height

    # Copiar celdas combinadas
    for merged_cell_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_cell_range))
        if min_row == max_row == source_row:
            new_range = (
                f"{ws.cell(row=target_row, column=min_col).coordinate}:"
                f"{ws.cell(row=target_row, column=max_col).coordinate}"
            )
            ws.merge_cells(new_range)


def export_entrance_requests_to_excel(
        db: Session, request_id: int, template_path: str, output_path: str):
    """Genera formato de ingreso a partir de una plantilla de Excel."""
    # Cargar datos de SQLAlchemy
    entrance_request = (
        db.query(EntranceRequest)
        .filter(EntranceRequest.id == request_id)
        .options(
            selectinload(EntranceRequest.branch),
            selectinload(EntranceRequest.branch).selectinload(Branch.municipality),
            selectinload(EntranceRequest.guests).selectinload(EntranceRequestGuest.guest),
            selectinload(
                EntranceRequest.guests
            ).selectinload(EntranceRequestGuest.guest).selectinload(Guest.eps),
            selectinload(
                EntranceRequest.guests
            ).selectinload(EntranceRequestGuest.guest).selectinload(Guest.arl),
            selectinload(
                EntranceRequest.guests
            ).selectinload(EntranceRequestGuest.guest).selectinload(Guest.company),
            selectinload(EntranceRequest.authorizer),
            selectinload(
                EntranceRequest.authorizer
            ).selectinload(User.unit),
            selectinload(
                EntranceRequest.authorizer
            ).selectinload(User.position),
            selectinload(EntranceRequest.creator),
            selectinload(
                EntranceRequest.creator
            ).selectinload(User.unit),
            selectinload(
                EntranceRequest.creator
            ).selectinload(User.position),
            selectinload(EntranceRequest.security),
            selectinload(
                EntranceRequest.security
            ).selectinload(User.unit),
            selectinload(
                EntranceRequest.security
            ).selectinload(User.position),
            selectinload(EntranceRequest.materials),
        )
        .first()
    )

    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Plantilla no encontrada: {template_path}")

    # Copiar plantilla a archivo nuevo
    shutil.copy(template_path, output_path)

    # Cargar archivo copiado
    wb = load_workbook(output_path)
    ws = wb.active

    # Datos Generales
    ws.cell(row=6, column=2).value = entrance_request.branch.name.upper()
    ws.cell(row=6, column=5).value = entrance_request.branch.municipality.name.upper()
    if entrance_request.branch.type == BranchTypes.administrative:
        ws.cell(row=5, column=9).value = ""
        ws.cell(row=6, column=9).value = "x"
    else:
        ws.cell(row=5, column=9).value = "x"
        ws.cell(row=6, column=9).value = ""
    ws.cell(row=5, column=11).value = "x" if entrance_request.is_installation else ""
    ws.cell(row=6, column=11).value = "x" if entrance_request.is_uninstallation else ""
    ws.cell(row=6, column=12).value = entrance_request.entry_date.strftime("%d/%m/%Y")
    ws.cell(row=6, column=15).value = entrance_request.departure_date.strftime("%d/%m/%Y")

    # Descripcion de las actividades
    ws.cell(row=9, column=2).value = entrance_request.reason.upper()

    # Solicitante, autorizador y seguridad
    ws.cell(row=28, column=3).value = entrance_request.creator.name
    ws.cell(row=28, column=8).value = entrance_request.authorizer.name
    ws.cell(row=28, column=15).value = entrance_request.security.name

    ws.cell(row=29, column=3).value = entrance_request.creator.unit.name
    ws.cell(row=29, column=8).value = entrance_request.authorizer.unit.name
    ws.cell(row=29, column=15).value = entrance_request.security.unit.name

    ws.cell(row=30, column=3).value = entrance_request.creator.position.name
    ws.cell(row=30, column=8).value = entrance_request.authorizer.position.name
    ws.cell(row=30, column=15).value = entrance_request.security.position.name

    ws.cell(row=31, column=3).value = entrance_request.creator.phone_number
    ws.cell(row=31, column=8).value = entrance_request.authorizer.phone_number
    ws.cell(row=31, column=15).value = entrance_request.security.phone_number

    # Relacion de ingreso y salida de personal a las instalaciones
    start_row = 15
    for idx, entrance_guest in enumerate(entrance_request.guests, start=start_row):
        if entrance_request.guests[-1] != entrance_guest:
            # Agregar una fila en blanco entre cada invitado
            current_row = idx
            ws.insert_rows(current_row)
            copy_row(ws, current_row, idx + 1)
        ws.cell(row=idx, column=2).value = entrance_guest.guest.name
        ws.cell(row=idx, column=4).value = entrance_guest.guest.eps.name
        ws.cell(row=idx, column=5).value = entrance_guest.guest.arl.name
        ws.cell(row=idx, column=7).value = entrance_guest.guest.document_id
        ws.cell(row=idx, column=8).value = entrance_guest.guest.company.name
        ws.cell(row=idx, column=14).value = entrance_request.entry_date.strftime("%H:%M")
        ws.cell(row=idx, column=16).value = entrance_request.departure_date.strftime("%H:%M")

    next_row = start_row + (len(entrance_request.guests) - 1) + 7
    # Inventario de materiales o equipos
    for idx, material in enumerate(entrance_request.materials, start=next_row):
        if entrance_request.materials[-1] != material:
            current_row = idx
            ws.insert_rows(current_row)
            copy_row(ws, current_row, idx + 1)
        ws.cell(row=idx, column=2).value = material.quantity
        ws.cell(row=idx, column=4).value = material.serial or ""
        ws.cell(row=idx, column=5).value = material.model
        ws.cell(row=idx, column=10).value = material.description or ""

    wb.save(output_path)
    print(f"Archivo generado: {output_path}")
